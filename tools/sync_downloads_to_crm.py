#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Метод Помощник — синк скачиваний в CRM + список для Meta.

Что делает:
  1. Читает экспорт таблицы скачиваний (CSV из Google-таблицы Apps Script).
  2. Собирает известные email из CRM и почты студентов из опросов.
  3. Тегирует каждого скачавшего: «студент» (есть в опросах) / «новый».
  4. Новых (нет в CRM) дописывает в лист «Сырые данные» CRM.
  5. Пишет meta_audience.csv — список email для загрузки в Meta Ads (Custom Audience).
  НЕ перезаписывает исходный CRM — сохраняет новый файл с суффиксом __synced.

Запуск:
  python3 tools/sync_downloads_to_crm.py <downloads.csv>

Как получить downloads.csv:
  Открой Google-таблицу скачиваний → Файл → Скачать → CSV. Положи путь в аргумент.
"""
import sys, os, csv, glob, re
from datetime import datetime
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CRM = os.path.join(ROOT, 'materials', 'CRM_Raw_Database_GAV.xlsx')
SURVEY_GLOB = os.path.join(ROOT, 'materials', '**', '*Ответы*.xlsx')
RAW_SHEET = 'Сырые данные'           # лист CRM для входящих лидов
RAW_HEADER = ['Имя','Фамилия','Позиция','Компания','Email','Телефон','Telegram',
              'Instagram','LinkedIn','Страна/Город','Источник','Продукт','Дата','Интерес']

EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

def norm(e):
    return (e or '').strip().lower()

def emails_from_xlsx(path):
    """Все валидные email из всех листов файла (ищем колонку с @)."""
    out = set()
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f'  ! не прочитал {os.path.basename(path)}: {e}'); return out
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and EMAIL_RE.match(v.strip()):
                    out.add(norm(v))
    return out

def main():
    if len(sys.argv) < 2:
        print('Использование: python3 tools/sync_downloads_to_crm.py <downloads.csv>')
        print('downloads.csv = экспорт Google-таблицы скачиваний (Файл → Скачать → CSV).')
        sys.exit(1)
    dl_path = sys.argv[1]
    if not os.path.exists(dl_path):
        print('Не найден файл:', dl_path); sys.exit(1)

    # 1. почты студентов (из опросов)
    student = set()
    for p in glob.glob(SURVEY_GLOB, recursive=True):
        student |= emails_from_xlsx(p)
    print(f'Почт студентов из опросов: {len(student)}')

    # 2. почты уже в CRM
    wb = openpyxl.load_workbook(CRM)
    crm_emails = set()
    for ws in wb.worksheets:
        try:
            hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
        except StopIteration:
            continue
        if 'Email' not in hdr: continue
        ei = hdr.index('Email')
        for row in ws.iter_rows(min_row=2, values_only=True):
            if ei < len(row) and isinstance(row[ei], str) and EMAIL_RE.match(row[ei].strip()):
                crm_emails.add(norm(row[ei]))
    print(f'Почт уже в CRM: {len(crm_emails)}')

    # 3. читаем скачивания
    with open(dl_path, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    def col(d, *names):
        for n in names:
            for k in d:
                if k and k.strip().lower() == n.lower(): return d[k]
        return ''
    seen, new_rows, audience, report = set(), [], [], {'студент':0,'новый':0,'всего':0,'дубль_в_файле':0}
    for r in rows:
        em = norm(col(r,'Email'))
        if not EMAIL_RE.match(em): continue
        report['всего'] += 1
        if em in seen: report['дубль_в_файле'] += 1; continue
        seen.add(em)
        audience.append(em)
        tag = 'студент' if em in student else 'новый'
        report[tag] += 1
        if em not in crm_emails:
            plat = col(r,'Платформа','platform')
            date = col(r,'Дата','Время (клиент)','ts')
            new_rows.append(['', '', '', '', col(r,'Email'), '', '', '', '', '',
                             'Метод Помощник (actas)', f'Method ({plat})' if plat else 'Method',
                             date or datetime.now().strftime('%Y-%m-%d'), tag])

    # 4. дописываем новых в «Сырые данные»
    if RAW_SHEET in wb.sheetnames:
        ws = wb[RAW_SHEET]
    else:
        ws = wb.create_sheet(RAW_SHEET); ws.append(RAW_HEADER)
    for nr in new_rows:
        ws.append(nr)
    out_crm = CRM.replace('.xlsx', '__synced.xlsx')
    wb.save(out_crm)

    # 5. список для Meta
    out_aud = os.path.join(ROOT, 'tools', 'meta_audience.csv')
    with open(out_aud, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f); w.writerow(['email'])
        for em in sorted(set(audience)): w.writerow([em])

    print('\n=== ИТОГ ===')
    print(f"Скачиваний в файле: {report['всего']} (дублей в файле: {report['дубль_в_файле']})")
    print(f"  студенты: {report['студент']} · новые: {report['новый']}")
    print(f"Новых лидов добавлено в CRM «Сырые данные»: {len(new_rows)}")
    print(f"CRM сохранён: {out_crm}  (исходный не тронут)")
    print(f"Список для Meta Ads: {out_aud}  ({len(set(audience))} email)")
    print("\nДальше: проверь __synced.xlsx, перенеси в основной CRM; meta_audience.csv → Ads Manager → Custom Audience (загрузка списка клиентов).")

if __name__ == '__main__':
    main()
